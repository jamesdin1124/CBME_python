import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = Workbook()

header_font = Font(name='Arial', bold=True, size=11, color='FFFFFF')
header_fill = PatternFill('solid', fgColor='2F5496')
sub_font = Font(name='Arial', bold=True, size=10)
sub_fill = PatternFill('solid', fgColor='D6E4F0')
nf = Font(name='Arial', size=10)
bf = Font(name='Arial', bold=True, size=10)
tf = Font(name='Arial', bold=True, size=14, color='2F5496')
tb = Border(left=Side('thin'), right=Side('thin'), top=Side('thin'), bottom=Side('thin'))
ca = Alignment(horizontal='center', vertical='center', wrap_text=True)
la = Alignment(horizontal='left', vertical='center', wrap_text=True)
input_fill = PatternFill('solid', fgColor='FFF2CC')
warn_fill = PatternFill('solid', fgColor='FFC7CE')

residents = ['游翔皓','呂昱暘','林盈秀','黃靖雯','李以琳','張元譯']

def style_hdr(ws, row, cols, font=None, fill=None):
    for c in range(1, cols+1):
        cell = ws.cell(row=row, column=c)
        cell.font = font or header_font
        cell.fill = fill or header_fill
        cell.alignment = ca
        cell.border = tb

def style_area(ws, r1, r2, cols):
    for r in range(r1, r2+1):
        for c in range(1, cols+1):
            cell = ws.cell(row=r, column=c)
            cell.font = nf
            cell.alignment = ca
            cell.border = tb

# Sheet 1: 基本資料
ws1 = wb.active
ws1.title = '基本資料'
ws1.sheet_properties.tabColor = '2F5496'
ws1['A1'] = '兒科住院醫師臨床表現評核表（主治醫師填寫）'
ws1['A1'].font = Font(name='Arial', bold=True, size=16, color='2F5496')
ws1.merge_cells('A1:H1')
ws1['A3'] = '評核期間：'; ws1['B3'] = '114 年 10 月 ～ 115 年 3 月（近半年）'
ws1['A4'] = '填表人（主治醫師）：'; ws1['B4'] = ''
ws1['A5'] = '填表日期：'; ws1['B5'] = ''
for r in range(3,6):
    ws1.cell(row=r, column=1).font = bf
    ws1.cell(row=r, column=2).font = nf
ws1.cell(row=4,column=2).fill = input_fill
ws1.cell(row=5,column=2).fill = input_fill

ws1['A7'] = '說明：請針對您近半年有共事經驗的住院醫師，就各面向給予評分。若無觀察機會請標註 N/A。'
ws1['A7'].font = nf
ws1.merge_cells('A7:H7')

ws1['A9'] = '住院醫師名單與訓練期程'
ws1['A9'].font = Font(name='Arial', bold=True, size=12, color='2F5496')
ws1.merge_cells('A9:H9')

for i, h in enumerate(['姓名','DOC碼','114-10','114-11','114-12','115-01','115-02','115-03'], 1):
    ws1.cell(row=10, column=i, value=h)
style_hdr(ws1, 10, 8)

rots = [
    ['游翔皓','DOC31030','外訓','PI','ER','PI','NB','ER'],
    ['呂昱暘','DOC31031','ER','外訓','外訓','外訓','PI','PI'],
    ['林盈秀','DOC32002','CN','ER','NB','CN','CN','NB'],
    ['黃靖雯','DOC32001','PI','NB','CN','NB','外訓','外訓'],
    ['李以琳','DOC32006','ER','CN/ER','NB','ER','W','NB'],
    ['張元譯','DOC32004','NB','NB/CN','PI','W','ER','CN'],
]
for ri, row in enumerate(rots, 11):
    for ci, val in enumerate(row, 1):
        ws1.cell(row=ri, column=ci, value=val)
style_area(ws1, 11, 16, 8)

ws1['A18'] = '站別代碼：ER=急診、NB=新生兒、PI=小兒加護、CN=一般兒科、W=病房、外訓=院外訓練'
ws1['A18'].font = Font(name='Arial', size=9, italic=True, color='666666')
ws1.merge_cells('A18:H18')

ws1.column_dimensions['A'].width = 14
ws1.column_dimensions['B'].width = 20
for col in 'CDEFGH':
    ws1.column_dimensions[col].width = 10

# Sheet 2: EPA評核
ws2 = wb.create_sheet('EPA評核')
ws2.sheet_properties.tabColor = '548235'
ws2['A1'] = '一、EPA 可信賴程度（Entrustable Professional Activities）'
ws2['A1'].font = tf
ws2.merge_cells('A1:G1')

ws2['A3'] = 'EPA 評分量表（9 級可信賴等級）'
ws2['A3'].font = Font(name='Arial', bold=True, size=11, color='2F5496')
ws2.merge_cells('A3:C3')

ws2.cell(row=4, column=1, value='文字等級')
ws2.cell(row=4, column=2, value='分數')
style_hdr(ws2, 4, 2, fill=PatternFill('solid', fgColor='548235'))

epa_scale = [
    ('允許住院醫師在旁觀察',1.5),('教師在旁逐步共同操作',2.0),('教師在旁必要時協助',2.5),
    ('教師可立即到場協助，事後逐項確認',3.0),('教師可立即到場協助，事後重點確認',3.3),
    ('教師可稍後到場協助，必要時事後確認',3.6),('教師 on call 提供監督',4.0),
    ('教師不需 on call，事後提供回饋及監督',4.5),('學員可對資淺學員進行監督與教學',5.0),
]
for ri, (txt, score) in enumerate(epa_scale, 5):
    ws2.cell(row=ri, column=1, value=txt)
    ws2.cell(row=ri, column=2, value=score)
style_area(ws2, 5, 13, 2)

ws2['A15'] = '≥ 3.5 → 熟練　｜　< 3.5 → 尚需加強'
ws2['A15'].font = Font(name='Arial', size=9, italic=True, color='C00000')
ws2.merge_cells('A15:C15')

ws2['A17'] = 'EPA 評核表（填入文字等級或對應分數 1.5–5.0）'
ws2['A17'].font = Font(name='Arial', bold=True, size=11, color='2F5496')
ws2.merge_cells('A17:G17')

for i, h in enumerate(['住院醫師','門診表現\n(OPD)','一般病人照護\n(WARD)','緊急處置\n(ED, DR)','重症照護\n(PICU, NICU)','病歷書寫','質性回饋'], 1):
    ws2.cell(row=18, column=i, value=h)
style_hdr(ws2, 18, 7)

for ri, name in enumerate(residents, 19):
    ws2.cell(row=ri, column=1, value=name)
    for ci in range(2, 8):
        ws2.cell(row=ri, column=ci).fill = input_fill
style_area(ws2, 19, 24, 7)

ws2.column_dimensions['A'].width = 12
for col in 'BCDEF':
    ws2.column_dimensions[col].width = 16
ws2.column_dimensions['G'].width = 25

# Sheet 3: 會議報告評核
ws3 = wb.create_sheet('會議報告評核')
ws3.sheet_properties.tabColor = 'BF8F00'
ws3['A1'] = '二、會議報告評核'
ws3['A1'].font = tf
ws3.merge_cells('A1:G1')

ws3['A3'] = '評分量表（5 級制）'
ws3['A3'].font = Font(name='Arial', bold=True, size=11, color='2F5496')
ws3.cell(row=4, column=1, value='文字等級')
ws3.cell(row=4, column=2, value='分數')
style_hdr(ws3, 4, 2, fill=PatternFill('solid', fgColor='BF8F00'))

for ri, (txt, score) in enumerate([('卓越',5),('充分',4),('尚可',3),('稍差',2),('不符合期待',1)], 5):
    ws3.cell(row=ri, column=1, value=txt)
    ws3.cell(row=ri, column=2, value=score)
style_area(ws3, 5, 9, 2)

ws3['A11'] = '適用會議：Staff Round、Journal Meeting、晨會指導、EBM 指導、多專科會議、MM'
ws3['A11'].font = Font(name='Arial', size=9, italic=True, color='666666')
ws3.merge_cells('A11:G11')

ws3['A13'] = '會議報告評核表（填入文字等級或對應分數 1–5）'
ws3['A13'].font = Font(name='Arial', bold=True, size=11, color='2F5496')
ws3.merge_cells('A13:G13')

for i, h in enumerate(['住院醫師','① 內容是否充分','② 辯證資料的能力','③ 口條呈現是否清晰','④ 開創建設性想法','⑤ 回答提問邏輯條理','質性回饋'], 1):
    ws3.cell(row=14, column=i, value=h)
style_hdr(ws3, 14, 7)

for ri, name in enumerate(residents, 15):
    ws3.cell(row=ri, column=1, value=name)
    for ci in range(2, 8):
        ws3.cell(row=ri, column=ci).fill = input_fill
style_area(ws3, 15, 20, 7)

ws3.column_dimensions['A'].width = 12
for col in 'BCDEF':
    ws3.column_dimensions[col].width = 18
ws3.column_dimensions['G'].width = 25

# Sheet 4: 操作技術-導管插管
ws4 = wb.create_sheet('操作技術-導管插管')
ws4.sheet_properties.tabColor = 'C00000'
ws4['A1'] = '三、操作技術評核 — 導管與插管類（8 項）'
ws4['A1'].font = tf
ws4.merge_cells('A1:J1')

ws4['A3'] = '使用 9 級可信賴等級（1.5–5.0），每格填寫：等級分數 / 觀察次數（例如 3.0 / 2次）'
ws4['A3'].font = Font(name='Arial', size=9, italic=True, color='666666')
ws4.merge_cells('A3:J3')

for i, h in enumerate(['住院醫師','插氣管內管\n(≥3)','插臍動靜脈\n導管 (≥1)','腰椎穿刺\n(≥3)','插 CVC\n(≥3)','肋膜液/腹水\n抽取 (≥1)','插胸管\n(≥2)','放置動脈\n導管 (≥2)','PICC\n(≥3)','質性回饋'], 1):
    ws4.cell(row=4, column=i, value=h)
style_hdr(ws4, 4, 10, fill=PatternFill('solid', fgColor='C00000'))

for ri, name in enumerate(residents, 5):
    ws4.cell(row=ri, column=1, value=name)
    for ci in range(2, 11):
        ws4.cell(row=ri, column=ci).fill = input_fill
style_area(ws4, 5, 10, 10)

ws4.column_dimensions['A'].width = 12
for col in 'BCDEFGHI':
    ws4.column_dimensions[col].width = 15
ws4.column_dimensions['J'].width = 25

# Sheet 5: 操作技術-超音波
ws5 = wb.create_sheet('操作技術-超音波')
ws5.sheet_properties.tabColor = '7030A0'
ws5['A1'] = '三、操作技術評核 — 超音波類（4 項）'
ws5['A1'].font = tf
ws5.merge_cells('A1:F1')

ws5['A3'] = '使用 9 級可信賴等級（1.5–5.0），每格填寫：等級分數 / 觀察次數'
ws5['A3'].font = Font(name='Arial', size=9, italic=True, color='666666')
ws5.merge_cells('A3:F3')

for i, h in enumerate(['住院醫師','腦部超音波\n(≥5)','心臟超音波\n(≥5)','腹部超音波\n(≥5)','腎臟超音波\n(≥5)','質性回饋'], 1):
    ws5.cell(row=4, column=i, value=h)
style_hdr(ws5, 4, 6, fill=PatternFill('solid', fgColor='7030A0'))

for ri, name in enumerate(residents, 5):
    ws5.cell(row=ri, column=1, value=name)
    for ci in range(2, 7):
        ws5.cell(row=ri, column=ci).fill = input_fill
style_area(ws5, 5, 10, 6)

ws5.column_dimensions['A'].width = 12
for col in 'BCDE':
    ws5.column_dimensions[col].width = 16
ws5.column_dimensions['F'].width = 25

# Sheet 6: CCC門檢標準
ws6 = wb.create_sheet('CCC門檢標準')
ws6.sheet_properties.tabColor = '0070C0'
ws6['A1'] = '四、CCC 門檢標準'
ws6['A1'].font = tf
ws6.merge_cells('A1:C1')

ws6['A3'] = '技能完成率門檻（依年級）'
ws6['A3'].font = Font(name='Arial', bold=True, size=11, color='2F5496')
for i, h in enumerate(['年級','通過標準','警示標準'], 1):
    ws6.cell(row=4, column=i, value=h)
style_hdr(ws6, 4, 3, fill=PatternFill('solid', fgColor='0070C0'))

for ri, (g, p, w) in enumerate([('R1','≥ 30% 項目達最低次數','< 30%'),('R2','≥ 60% 項目達最低次數','< 60%'),('R3','≥ 90% 項目達最低次數','< 90%')], 5):
    ws6.cell(row=ri, column=1, value=g)
    ws6.cell(row=ri, column=2, value=p)
    ws6.cell(row=ri, column=3, value=w)
style_area(ws6, 5, 7, 3)
for r in range(5,8):
    ws6.cell(row=r, column=3).fill = warn_fill

ws6['A9'] = 'EPA / 會議報告分數門檻'
ws6['A9'].font = Font(name='Arial', bold=True, size=11, color='2F5496')
for i, h in enumerate(['面向','通過','警示'], 1):
    ws6.cell(row=10, column=i, value=h)
style_hdr(ws6, 10, 3, fill=PatternFill('solid', fgColor='0070C0'))

ws6.cell(row=11, column=1, value='EPA 平均分')
ws6.cell(row=11, column=2, value='≥ 3.5')
ws6.cell(row=11, column=3, value='< 2.5')
ws6.cell(row=12, column=1, value='會議報告平均分')
ws6.cell(row=12, column=2, value='≥ 3.5')
ws6.cell(row=12, column=3, value='< 2.5')
style_area(ws6, 11, 12, 3)
for r in (11,12):
    ws6.cell(row=r, column=3).fill = warn_fill

ws6.column_dimensions['A'].width = 16
ws6.column_dimensions['B'].width = 25
ws6.column_dimensions['C'].width = 15

# Sheet 7: 補充意見
ws7 = wb.create_sheet('補充意見')
ws7.sheet_properties.tabColor = '00B050'
ws7['A1'] = '五、補充意見'
ws7['A1'].font = tf
ws7.merge_cells('A1:D1')

ws7['A3'] = '特別表現優異之住院醫師與具體事蹟：'
ws7['A3'].font = bf
ws7.merge_cells('A4:D8')
ws7['A4'].fill = input_fill
ws7['A4'].alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
for r in range(4,9):
    for c in range(1,5):
        ws7.cell(row=r, column=c).border = tb

ws7['A10'] = '需特別關注或輔導之住院醫師與具體事項：'
ws7['A10'].font = bf
ws7.merge_cells('A11:D15')
ws7['A11'].fill = input_fill
ws7['A11'].alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
for r in range(11,16):
    for c in range(1,5):
        ws7.cell(row=r, column=c).border = tb

ws7['A17'] = '對訓練計畫之建議：'
ws7['A17'].font = bf
ws7.merge_cells('A18:D22')
ws7['A18'].fill = input_fill
ws7['A18'].alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
for r in range(18,23):
    for c in range(1,5):
        ws7.cell(row=r, column=c).border = tb

for col in 'ABCD':
    ws7.column_dimensions[col].width = 20

for ws in wb.worksheets:
    ws.sheet_format.defaultRowHeight = 20
    ws.page_setup.orientation = 'landscape'
    ws.page_setup.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0

import os
output = os.path.join(os.path.dirname(os.path.abspath(__file__)), '兒科住院醫師評核表.xlsx')
wb.save(output)
print(f'Saved to {output}')
