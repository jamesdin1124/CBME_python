"""
兒科住院醫師自我評核表 — 產生器
用法：python3 create_self_eval.py
產出：住院醫師自評表.xlsx（同目錄下）
"""
import openpyxl, os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

wb = Workbook()

header_font = Font(name='Arial', bold=True, size=11, color='FFFFFF')
header_fill = PatternFill('solid', fgColor='2F5496')
nf = Font(name='Arial', size=10)
bf = Font(name='Arial', bold=True, size=10)
tf = Font(name='Arial', bold=True, size=14, color='2F5496')
tb = Border(left=Side('thin'), right=Side('thin'), top=Side('thin'), bottom=Side('thin'))
ca = Alignment(horizontal='center', vertical='center', wrap_text=True)
la = Alignment(horizontal='left', vertical='center', wrap_text=True)
la_top = Alignment(horizontal='left', vertical='top', wrap_text=True)
input_fill = PatternFill('solid', fgColor='FFF2CC')
light_blue = PatternFill('solid', fgColor='D6E4F0')
light_purple = PatternFill('solid', fgColor='E8D5F5')

def style_hdr(ws, row, cols, font=None, fill=None):
    for c in range(1, cols+1):
        cell = ws.cell(row=row, column=c)
        cell.font = font or header_font
        cell.fill = fill or header_fill
        cell.alignment = ca
        cell.border = tb

def style_area(ws, r1, r2, cols, fill_from=None):
    for r in range(r1, r2+1):
        for c in range(1, cols+1):
            cell = ws.cell(row=r, column=c)
            cell.font = nf
            cell.alignment = ca
            cell.border = tb

def add_input_row(ws, row, label, col_label, col_input, merge_to=None):
    ws.cell(row=row, column=col_label, value=label).font = bf
    ws.cell(row=row, column=col_label).alignment = la
    ws.cell(row=row, column=col_label).border = tb
    inp = ws.cell(row=row, column=col_input)
    inp.fill = input_fill
    inp.border = tb
    inp.alignment = la
    if merge_to:
        ws.merge_cells(start_row=row, start_column=col_input, end_row=row, end_column=merge_to)
        for c in range(col_input, merge_to+1):
            ws.cell(row=row, column=c).border = tb

# ==========================================
# Sheet 1: 自我評核（EPA + 會議報告）
# ==========================================
ws1 = wb.active
ws1.title = '自我評核'
ws1.sheet_properties.tabColor = '2F5496'

ws1['A1'] = '兒科住院醫師自我評核表'
ws1['A1'].font = Font(name='Arial', bold=True, size=16, color='2F5496')
ws1.merge_cells('A1:G1')

ws1['A2'] = '（住院醫師自填，請如實填寫各項目）'
ws1['A2'].font = Font(name='Arial', size=10, italic=True, color='666666')
ws1.merge_cells('A2:G2')

# ----- 基本資料 -----
row = 4
add_input_row(ws1, row, '住院醫師姓名：', 1, 2, 3)
add_input_row(ws1, row, '訓練年級：', 4, 5, 5)
ws1.cell(row=row, column=6, value='（R1 / R2 / R3）').font = Font(name='Arial', size=9, italic=True, color='999999')
row = 5
add_input_row(ws1, row, '評核期間：', 1, 2, 3)
add_input_row(ws1, row, '科別：', 4, 5, 5)
row = 6
add_input_row(ws1, row, '填表日期：', 1, 2, 3)

# ----- EPA 自評 -----
row = 8
ws1.cell(row=row, column=1, value='一、EPA 可信賴程度自評').font = Font(name='Arial', bold=True, size=12, color='2F5496')
ws1.merge_cells(f'A{row}:G{row}')

row = 9
ws1.cell(row=row, column=1, value='請參照 9 級可信賴等級，自評您目前在各面向的能力等級（1.5–5.0）').font = Font(name='Arial', size=9, italic=True, color='666666')
ws1.merge_cells(f'A{row}:G{row}')

row = 11
ws1.cell(row=row, column=1, value='EPA 評分量表（9 級可信賴等級）').font = Font(name='Arial', bold=True, size=10, color='548235')
ws1.merge_cells(f'A{row}:C{row}')

row = 12
ws1.cell(row=row, column=1, value='文字等級')
ws1.cell(row=row, column=2, value='分數')
style_hdr(ws1, row, 2, fill=PatternFill('solid', fgColor='548235'))

epa_scale = [
    ('允許住院醫師在旁觀察', 1.5), ('教師在旁逐步共同操作', 2.0), ('教師在旁必要時協助', 2.5),
    ('教師可立即到場協助，事後逐項確認', 3.0), ('教師可立即到場協助，事後重點確認', 3.3),
    ('教師可稍後到場協助，必要時事後確認', 3.6), ('教師 on call 提供監督', 4.0),
    ('教師不需 on call，事後提供回饋及監督', 4.5), ('學員可對資淺學員進行監督與教學', 5.0),
]
for ri, (txt, score) in enumerate(epa_scale, 13):
    ws1.cell(row=ri, column=1, value=txt).font = nf
    ws1.cell(row=ri, column=1).alignment = la
    ws1.cell(row=ri, column=1).border = tb
    ws1.cell(row=ri, column=2, value=score).font = nf
    ws1.cell(row=ri, column=2).alignment = ca
    ws1.cell(row=ri, column=2).border = tb

row = 22
ws1.cell(row=row, column=1, value='≥ 3.5 → 熟練　｜　< 3.5 → 尚需加強').font = Font(name='Arial', size=9, italic=True, color='C00000')
ws1.merge_cells(f'A{row}:C{row}')

row = 24
ws1.cell(row=row, column=1, value='EPA 自我評核').font = Font(name='Arial', bold=True, size=10, color='2F5496')
ws1.merge_cells(f'A{row}:D{row}')

row = 25
for i, h in enumerate(['評核面向', '自評分數\n(1.5–5.0)', '自評文字等級', '自覺優勢 / 待加強之處'], 1):
    ws1.cell(row=row, column=i, value=h)
style_hdr(ws1, row, 4, fill=PatternFill('solid', fgColor='548235'))

for ri, item in enumerate(['門診表現 (OPD)', '一般病人照護 (WARD)', '緊急處置 (ED, DR)', '重症照護 (PICU, NICU)', '病歷書寫'], 26):
    ws1.cell(row=ri, column=1, value=item).font = nf
    ws1.cell(row=ri, column=1).alignment = la
    ws1.cell(row=ri, column=1).border = tb
    for ci in range(2, 5):
        ws1.cell(row=ri, column=ci).fill = input_fill
        ws1.cell(row=ri, column=ci).border = tb
        ws1.cell(row=ri, column=ci).alignment = ca if ci <= 3 else la

# ----- 會議報告自評 -----
row = 33
ws1.cell(row=row, column=1, value='二、會議報告自評').font = Font(name='Arial', bold=True, size=12, color='2F5496')
ws1.merge_cells(f'A{row}:G{row}')

row = 34
ws1.cell(row=row, column=1, value='5 級制：卓越(5) / 充分(4) / 尚可(3) / 稍差(2) / 不符合期待(1)').font = Font(name='Arial', size=9, italic=True, color='666666')
ws1.merge_cells(f'A{row}:G{row}')

row = 36
for i, h in enumerate(['評核面向', '自評分數\n(1–5)', '自覺優勢 / 待加強之處'], 1):
    ws1.cell(row=row, column=i, value=h)
style_hdr(ws1, row, 3, fill=PatternFill('solid', fgColor='BF8F00'))

for ri, item in enumerate(['① 內容是否充分', '② 辯證資料的能力', '③ 口條呈現是否清晰', '④ 開創建設性想法', '⑤ 回答提問邏輯條理'], 37):
    ws1.cell(row=ri, column=1, value=item).font = nf
    ws1.cell(row=ri, column=1).alignment = la
    ws1.cell(row=ri, column=1).border = tb
    for ci in range(2, 4):
        ws1.cell(row=ri, column=ci).fill = input_fill
        ws1.cell(row=ri, column=ci).border = tb
        ws1.cell(row=ri, column=ci).alignment = ca if ci == 2 else la

ws1.column_dimensions['A'].width = 30
ws1.column_dimensions['B'].width = 15
ws1.column_dimensions['C'].width = 18
ws1.column_dimensions['D'].width = 35
ws1.column_dimensions['E'].width = 15
ws1.column_dimensions['F'].width = 15
ws1.column_dimensions['G'].width = 15

# ==========================================
# Sheet 2: Procedure 紀錄（詳細 Log）
# ==========================================
ws2 = wb.create_sheet('Procedure紀錄')
ws2.sheet_properties.tabColor = 'C00000'

ws2['A1'] = '三、操作技術紀錄（Procedure Log）'
ws2['A1'].font = tf
ws2.merge_cells('A1:H1')

ws2['A2'] = '請詳列您在本評核期間內實際執行過的 procedure，並填寫病歷號以供查核。每項 procedure 可填寫多次。'
ws2['A2'].font = Font(name='Arial', size=10, italic=True, color='666666')
ws2.merge_cells('A2:H2')

proc_headers = ['序號', 'Procedure 名稱', '執行日期', '病歷號', '病人年齡', '執行角色\n(主執行/協助)', '督導醫師', '備註']

# --- A. 導管與插管類 ---
row = 4
ws2.cell(row=row, column=1, value='A. 導管與插管類').font = Font(name='Arial', bold=True, size=11, color='C00000')
ws2.merge_cells(f'A{row}:H{row}')

row = 5
for i, h in enumerate(proc_headers, 1):
    ws2.cell(row=row, column=i, value=h)
style_hdr(ws2, row, 8, fill=PatternFill('solid', fgColor='C00000'))

catheter_procs = [
    '插氣管內管 (Intubation)', '插臍動靜脈導管 (UAC/UVC)', '腰椎穿刺 (LP)',
    '插 CVC (Central Line)', '肋膜液/腹水抽取 (Thoracentesis/Paracentesis)',
    '插胸管 (Chest Tube)', '放置動脈導管 (Arterial Line)', 'PICC',
]

r = 6
for idx, proc in enumerate(catheter_procs, 1):
    for extra in range(3):  # 3 rows per procedure
        ws2.cell(row=r, column=1, value=idx if extra == 0 else '').font = nf
        ws2.cell(row=r, column=1).alignment = ca
        ws2.cell(row=r, column=1).border = tb
        ws2.cell(row=r, column=2, value=proc if extra == 0 else '').font = nf
        ws2.cell(row=r, column=2).alignment = la
        ws2.cell(row=r, column=2).border = tb
        ws2.cell(row=r, column=2).fill = light_blue
        for ci in range(3, 9):
            ws2.cell(row=r, column=ci).fill = input_fill
            ws2.cell(row=r, column=ci).border = tb
            ws2.cell(row=r, column=ci).alignment = ca
        r += 1

# --- B. 超音波類 ---
r += 1
ws2.cell(row=r, column=1, value='B. 超音波類').font = Font(name='Arial', bold=True, size=11, color='7030A0')
ws2.merge_cells(f'A{r}:H{r}')
r += 1
for i, h in enumerate(proc_headers, 1):
    ws2.cell(row=r, column=i, value=h)
style_hdr(ws2, r, 8, fill=PatternFill('solid', fgColor='7030A0'))
r += 1

echo_procs = ['腦部超音波 (Brain Echo)', '心臟超音波 (Echo)', '腹部超音波 (Abdomen Echo)', '腎臟超音波 (Renal Echo)']
for idx, proc in enumerate(echo_procs, 1):
    for extra in range(3):
        ws2.cell(row=r, column=1, value=idx if extra == 0 else '').font = nf
        ws2.cell(row=r, column=1).alignment = ca
        ws2.cell(row=r, column=1).border = tb
        ws2.cell(row=r, column=2, value=proc if extra == 0 else '').font = nf
        ws2.cell(row=r, column=2).alignment = la
        ws2.cell(row=r, column=2).border = tb
        ws2.cell(row=r, column=2).fill = light_purple
        for ci in range(3, 9):
            ws2.cell(row=r, column=ci).fill = input_fill
            ws2.cell(row=r, column=ci).border = tb
            ws2.cell(row=r, column=ci).alignment = ca
        r += 1

# --- C. 其他 Procedure ---
r += 1
ws2.cell(row=r, column=1, value='C. 其他 Procedure（自行填寫）').font = Font(name='Arial', bold=True, size=11, color='00B050')
ws2.merge_cells(f'A{r}:H{r}')
r += 1
for i, h in enumerate(proc_headers, 1):
    ws2.cell(row=r, column=i, value=h)
style_hdr(ws2, r, 8, fill=PatternFill('solid', fgColor='00B050'))
r += 1

for idx in range(1, 16):
    ws2.cell(row=r, column=1, value=idx).font = nf
    ws2.cell(row=r, column=1).alignment = ca
    ws2.cell(row=r, column=1).border = tb
    for ci in range(2, 9):
        ws2.cell(row=r, column=ci).fill = input_fill
        ws2.cell(row=r, column=ci).border = tb
        ws2.cell(row=r, column=ci).alignment = ca if ci != 2 else la
    r += 1

ws2.column_dimensions['A'].width = 8
ws2.column_dimensions['B'].width = 32
ws2.column_dimensions['C'].width = 14
ws2.column_dimensions['D'].width = 14
ws2.column_dimensions['E'].width = 12
ws2.column_dimensions['F'].width = 16
ws2.column_dimensions['G'].width = 14
ws2.column_dimensions['H'].width = 20

# ==========================================
# Sheet 3: Procedure 統計摘要
# ==========================================
ws3 = wb.create_sheet('Procedure統計')
ws3.sheet_properties.tabColor = 'FF6600'

ws3['A1'] = '操作技術統計摘要'
ws3['A1'].font = tf
ws3.merge_cells('A1:E1')

ws3['A2'] = '請統計您在本評核期間各 procedure 的總執行次數與自評可信賴等級'
ws3['A2'].font = Font(name='Arial', size=10, italic=True, color='666666')
ws3.merge_cells('A2:E2')

row = 4
for i, h in enumerate(['Procedure 名稱', '最低要求次數', '實際完成次數', '自評可信賴等級\n(1.5–5.0)', '是否達標\n(是/否)'], 1):
    ws3.cell(row=row, column=i, value=h)
style_hdr(ws3, row, 5, fill=PatternFill('solid', fgColor='FF6600'))

all_procs = [
    ('插氣管內管', '≥3'), ('插臍動靜脈導管', '≥1'), ('腰椎穿刺', '≥3'),
    ('插 CVC', '≥3'), ('肋膜液/腹水抽取', '≥1'), ('插胸管', '≥2'),
    ('放置動脈導管', '≥2'), ('PICC', '≥3'),
    ('腦部超音波', '≥5'), ('心臟超音波', '≥5'), ('腹部超音波', '≥5'), ('腎臟超音波', '≥5'),
]
row = 5
for proc_name, min_req in all_procs:
    ws3.cell(row=row, column=1, value=proc_name).font = nf
    ws3.cell(row=row, column=1).alignment = la
    ws3.cell(row=row, column=1).border = tb
    ws3.cell(row=row, column=2, value=min_req).font = nf
    ws3.cell(row=row, column=2).alignment = ca
    ws3.cell(row=row, column=2).border = tb
    for ci in range(3, 6):
        ws3.cell(row=row, column=ci).fill = input_fill
        ws3.cell(row=row, column=ci).border = tb
        ws3.cell(row=row, column=ci).alignment = ca
    row += 1

ws3.column_dimensions['A'].width = 22
ws3.column_dimensions['B'].width = 16
ws3.column_dimensions['C'].width = 16
ws3.column_dimensions['D'].width = 20
ws3.column_dimensions['E'].width = 14

# ==========================================
# Sheet 4: 自我反思
# ==========================================
ws4 = wb.create_sheet('自我反思')
ws4.sheet_properties.tabColor = '00B050'

ws4['A1'] = '四、自我反思與學習計畫'
ws4['A1'].font = tf
ws4.merge_cells('A1:D1')

row = 3
for item in [
    '本期間自覺最大的進步或成就：',
    '本期間遇到最大的困難或挑戰：',
    '目前最需要加強的臨床能力：',
    '下一期的學習目標與計畫：',
    '希望主治醫師 / 科部提供的協助或資源：',
]:
    ws4.cell(row=row, column=1, value=item).font = bf
    ws4.cell(row=row, column=1).alignment = la
    ws4.merge_cells(f'A{row}:D{row}')
    row += 1
    ws4.merge_cells(f'A{row}:D{row+3}')
    ws4.cell(row=row, column=1).fill = input_fill
    ws4.cell(row=row, column=1).alignment = la_top
    for rr in range(row, row+4):
        for cc in range(1, 5):
            ws4.cell(row=rr, column=cc).border = tb
    row += 5

for col in 'ABCD':
    ws4.column_dimensions[col].width = 22

# ==========================================
# Global page settings
# ==========================================
for ws in wb.worksheets:
    ws.sheet_format.defaultRowHeight = 20
    ws.page_setup.orientation = 'landscape'
    ws.page_setup.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0

output = os.path.join(os.path.dirname(os.path.abspath(__file__)), '住院醫師自評表.xlsx')
wb.save(output)
print(f'✅ 已產生：{output}')
